import os
import time
import logging
from typing import Literal
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from pywinauto import Application, Desktop
from openpyxl.utils import coordinate_to_tuple
from pywinauto.timings import wait_until_passes
from openpyxl.worksheet.worksheet import Worksheet
from pywinauto.controls.uiawrapper import UIAWrapper
from src.common.decorator import HandleExceptionMethod
from pywinauto.application import WindowSpecification
from pywinauto.controls.uia_controls import ButtonWrapper


class Excel:
    @HandleExceptionMethod()
    def __init__(
        self,
        file_path: str,
        timeout: int = 10,
        retry_interval: float = 0.5,
        auto_save: bool = False,
        logger_name: str = __name__,
    ):
        self.logger = logging.getLogger(logger_name)
        if not os.path.exists(file_path):
            self.logger.error(f"{file_path} không tồn tại")
            raise FileNotFoundError(f"{file_path} không tồn tại")
        self.file_path = file_path
        self.timeout = timeout
        self.auto_save = auto_save
        self.retry_interval = retry_interval
        cmd = r'"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE" "{}"'.format(
                file_path
        )
        print(cmd)
        Application(backend="uia").start(cmd)
        while True:
            titles = [w.window_text() for w in Desktop(backend="uia").windows()]
            if found := [
                title
                for title in titles
                if title.startswith(os.path.basename(self.file_path))
            ]:
                self.title: str = found[0]
                break
            continue
        self.App: Application = Application(backend="uia").connect(title=self.title)
        if self.title.find("[Protected View]") != -1:
            ExcelApp: WindowSpecification = self.App.window(title=self.title)
            EnableEditingButton: UIAWrapper = wait_until_passes(
                timeout=self.timeout,
                retry_interval=self.retry_interval,
                func=lambda: ExcelApp.child_window(
                    title="Enable Editing", control_type="Button"
                ).wrapper_object(),
            )
            EnableEditingButton.click_input(double=True)
        time.sleep(5)
        if self.App.window().window_text().find("[Protected View]") != -1:
            return self.__init__(
                file_path=self.file_path,
                timeout=self.timeout,
                retry_interval=self.retry_interval,
                auto_save=self.auto_save,
                logger_name=self.logger.name,
            )
        else:
            self.title = self.App.window().window_text()

    @HandleExceptionMethod()
    def __del__(self):
        ExcelApp: WindowSpecification = self.App.window(title=self.title)
        ExcelApp.close()
        time.sleep(2.5)
        if not self.auto_save:
            wait_until_passes(
                timeout=self.timeout,
                retry_interval=self.retry_interval,
                func=lambda: ExcelApp.child_window(
                    title="Don't Save", control_type="Button"
                ).exists(),
            )
            DontSaveButton: UIAWrapper = ExcelApp.child_window(
                title="Don't Save", control_type="Button"
            ).wrapper_object()
            DontSaveButton.click_input()
        else:
            wait_until_passes(
                timeout=self.timeout,
                retry_interval=self.retry_interval,
                func=lambda: ExcelApp.child_window(
                    title="Save", control_type="Button"
                ).exists(),
            )
            DontSaveButton: UIAWrapper = ExcelApp.child_window(
                title="Save", control_type="Button"
            ).wrapper_object()
            DontSaveButton.click_input()

    @HandleExceptionMethod()
    def __wait_for_exists(self, window: WindowSpecification) -> UIAWrapper | None:
        wait_until_passes(
            timeout=self.timeout,
            retry_interval=self.retry_interval,
            func=lambda: window.exists(timeout=self.timeout),
        )
        return window.wrapper_object()

    @property
    def shape(self, sheet: str = "Sheet1") -> tuple[str]:
        workbook = load_workbook(self.file_path)
        sheet = workbook[sheet]
        return tuple(sheet.dimensions.split(":"))

    @HandleExceptionMethod()
    def page_setup(
        self,
        orientation: Literal["Landscape", "Portrait"] = "Portrait",
        header: str = None,
    ) -> bool:
        self.logger.info(f"Page Setup {self.title}")
        ExcelApp: WindowSpecification = self.App.window(title=self.title)
        self.__wait_for_exists(
            ExcelApp.child_window(
                title="Page Layout",
                auto_id="TabPageLayoutExcel",
                control_type="TabItem",
            )
        ).select()
        self.__wait_for_exists(
            ExcelApp.child_window(
                title="Print Titles", auto_id="PrintTitles", control_type="Button"
            )
        ).click_input(double=True)
        while not self.App.window(title="Page Setup").exists():
            continue
        dialog = self.App.window(title="Page Setup")

        self.__wait_for_exists(
            dialog.child_window(title="Page", control_type="TabItem")
        ).select()
        self.__wait_for_exists(
            dialog.child_window(title=orientation, control_type="RadioButton")
        ).click_input()
        #
        if header:
            self.__wait_for_exists(
                dialog.child_window(title="Header/Footer", control_type="TabItem")
            ).select()
            try:
                self.__wait_for_exists(
                    dialog.child_window(title="Header:", control_type="ComboBox")
                ).select(header).click_input()
            except IndexError:
                self.__wait_for_exists(
                    dialog.child_window(title="Header/Footer", control_type="TabItem")
                ).select()
                self.__wait_for_exists(
                    dialog.child_window(title="Custom Header...", control_type="Button")
                ).click_input(double=True)
                self.__wait_for_exists(
                    dialog.child_window(title="Left section:", control_type="Edit")
                ).type_keys(header)
                buttons: list[ButtonWrapper] = dialog.descendants(
                    title="OK", control_type="Button"
                )
                button = [
                    button
                    for button in buttons
                    if button.parent().window_text() == "Header"
                ][0]
                button.click_input()
                time.sleep(1)
        self.__wait_for_exists(
            dialog.child_window(title="OK", control_type="Button")
        ).click_input()
        return True

    @HandleExceptionMethod()
    def edit(
        self,
        cells: list[str] = [],
        contents: list[str] = [],
        background_colors: list[str] = [],
        sheet: str = "Sheet1",
    ) -> bool:
        # self.logger.info(f"{sheet}: {cell}='{content}'")
        # Exit App #
        self.__del__()
        time.sleep(2.5)
        workbook = load_workbook(self.file_path)
        sheet: Worksheet = workbook[sheet]
        # Edit
        for index, cell in enumerate(cells):
            row, col = coordinate_to_tuple(cell)
            content = contents[index] if len(contents) >= index + 1 else None
            background_color = (
                background_colors[index]
                if len(background_colors) >= index + 1
                else None
            )
            self.logger.info(
                f"{sheet}, {cell}(background={background_color}) = {content}"
            )
            if content:
                cell: Cell = sheet.cell(row=row, column=col, value=content)
            if background_color:
                cell: Cell = sheet.cell(row=row, column=col)
                fill = PatternFill(fill_type="solid", fgColor=background_color)
                cell.fill = fill
        workbook.save(self.file_path)
        # Reopen
        self.__init__(
            file_path=self.file_path,
            timeout=self.timeout,
            retry_interval=self.retry_interval,
        )
        return True

    @HandleExceptionMethod()
    def save(self):
        ExcelApp: WindowSpecification = self.App.window(title=self.title)
        SaveButton = self.__wait_for_exists(
            ExcelApp.child_window(
                title="Save", auto_id="FileSave", control_type="Button"
            )
        )
        SaveButton.click_input()
        SaveButton.click_input()
        SaveButton.click_input()

    @HandleExceptionMethod()
    def format_cells(self,range:str,tab:str="Number",**kwargs):
        ExcelApp: WindowSpecification = self.App.window(title=self.title)
        self.__wait_for_exists(
            ExcelApp.child_window(auto_id="A1", control_type="DataItem")
        ).type_keys("{ENTER}")
        NameBox = self.__wait_for_exists(
            ExcelApp.child_window(title="Name Box", auto_id="1001", control_type="Edit")
        )
        NameBox.click_input(double=True)
        NameBox.type_keys(range)
        NameBox.type_keys("{ENTER}", pause=1)
        self.__wait_for_exists(
            ExcelApp.child_window(
                title="Home", auto_id="TabHome", control_type="TabItem"
            )
        ).select()
        FormatCellsMenu = self.__wait_for_exists(
            ExcelApp.child_window(
                title="Format", auto_id="FormatCellsMenu", control_type="MenuItem"
            )
        )
        FormatCellsMenu.click_input()
        FormatCellsMenu.click_input(
            button_down=False,
        )
        FormatCellNumber = self.__wait_for_exists(
            ExcelApp.child_window(title="Format Cell Number", auto_id="FormatCellsNumberDialog", control_type="Button")
        )
        FormatCellNumber.click_input(double=True)
        self.__wait_for_exists(
            ExcelApp.child_window(title=tab, control_type="TabItem")
        ).select()       
        self.__wait_for_exists(
            ExcelApp.child_window(title="Number", control_type="ListItem")
        ).select()
        Use1000Separator = self.__wait_for_exists(
            ExcelApp.child_window(title="Use 1000 Separator (.)", control_type="CheckBox")
        )
        while Use1000Separator.get_toggle_state() == 0:
            Use1000Separator.click_input()
        #-#
        import io
        import sys
        output = io.StringIO()
        sys.stdout = output
        ExcelApp.print_control_identifiers()
        sys.stdout = sys.__stdout__
        control_identifiers = output.getvalue()
        with open('control.txt', 'w', encoding='utf-8') as f:
            f.write(control_identifiers)
        #-#
        DecimalPlaces: UIAWrapper = self.__wait_for_exists(
            ExcelApp.child_window(title="Decimal places:", control_type="Spinner")
        )
        DecimalPlaces.click_input(double=True)
        DecimalPlaces.type_keys("^a{DELETE}",pause=1)
        DecimalPlaces.type_keys("0",pause=1)
        self.__wait_for_exists(
            ExcelApp.child_window(title="OK", control_type="Button")
        ).click_input(double=True)
    @HandleExceptionMethod()
    def format(
        self,
        AutoFitColumnWith: bool = False,
        Border: str = "Bottom Border",
    ):
        self.logger.info(f"Format {self.title}")
        ExcelApp: WindowSpecification = self.App.window(title=self.title)
        self.__wait_for_exists(
            ExcelApp.child_window(auto_id="A1", control_type="DataItem")
        ).type_keys("{ENTER}")
        NameBox = self.__wait_for_exists(
            ExcelApp.child_window(title="Name Box", auto_id="1001", control_type="Edit")
        )
        NameBox.click_input(double=True)
        NameBox.type_keys(f"{self.shape[0]}:{self.shape[1]}")
        NameBox.type_keys("{ENTER}", pause=1)
        # - AutoFitColumnWith - #
        if AutoFitColumnWith:
            self.__wait_for_exists(
                ExcelApp.child_window(
                    title="Home", auto_id="TabHome", control_type="TabItem"
                )
            ).select()
            FormatCellsMenu = self.__wait_for_exists(
                ExcelApp.child_window(
                    title="Format", auto_id="FormatCellsMenu", control_type="MenuItem"
                )
            )
            FormatCellsMenu.click_input()
            FormatCellsMenu.click_input(
                button_down=False,
            )
            self.__wait_for_exists(
                ExcelApp.child_window(
                    title="AutoFit Column Width", control_type="MenuItem"
                )
            ).click_input()
        # - Border - #
        self.__wait_for_exists(
            ExcelApp.child_window(
                title="More Options",
                auto_id="BordersGallery_Dropdown",
                control_type="MenuItem",
            )
        ).click_input()
        self.__wait_for_exists(
            ExcelApp.child_window(title=Border, control_type="MenuItem")
        ).click_input()
        # --#

    @HandleExceptionMethod()
    def export(
        self,
        file_name: str = None,
    ) -> str:
        if file_name and not file_name.endswith(".pdf"):
            self.logger.error("Chỉ hỗ trợ export sang PDF")
            return None
        self.logger.info(f"Export {self.title}")
        ExcelApp: WindowSpecification = self.App.window(title=self.title)
        self.__wait_for_exists(
            ExcelApp.child_window(
                title="File Tab", auto_id="FileTabButton", control_type="Button"
            )
        ).click_input()
        self.__wait_for_exists(
            ExcelApp.child_window(title="Export", control_type="ListItem")
        ).click_input()
        self.__wait_for_exists(
            ExcelApp.child_window(title="Create PDF/XPS", control_type="Button")
        ).click_input()
        time.sleep(2)
        CheckBox = self.__wait_for_exists(
            ExcelApp.child_window(
                title="Open file after publishing", control_type="CheckBox"
            )
        )
        if CheckBox.get_toggle_state() == 1:
            CheckBox.click_input()
        if file_name:
            self.__wait_for_exists(
                ExcelApp.child_window(
                    title="File name:",
                    auto_id="FileNameControlHost",
                    control_type="ComboBox",
                )
            ).type_keys(file_name, pause=0.5)
        self.__wait_for_exists(
            ExcelApp.child_window(title="Publish", auto_id="1", control_type="Button")
        ).click_input()
        time.sleep(0.5)
        if ExcelApp.child_window(
            title="Confirm Save As", control_type="Window"
        ).exists():
            self.__wait_for_exists(
                ExcelApp.child_window(
                    title="Yes", auto_id="CommandButton_6", control_type="Button"
                )
            ).click_input()
        if file_name:
            return file_name
        else:
            base, _ = os.path.splitext(self.file_path)
            return base + ".pdf"

    @HandleExceptionMethod()
    def search(
        self, keyword: str, axis: int = 0, sheetname: str = "Sheet1"
    ) -> str | None:
        workbook = load_workbook(
            filename=self.file_path,
            data_only=True,
        )
        worksheet = workbook[sheetname]
        target_cell: Cell | None = None
        cells: tuple[Cell, ...] = None
        if axis == 0:  # Tìm kiếm theo cột dọc (column)
            for col in worksheet.iter_cols():
                for cell in col:
                    if cell.value == keyword:
                        target_cell = cell
                        break
                if target_cell:
                    cells = [
                        row[0]
                        for row in worksheet.iter_rows(
                            min_col=target_cell.column, max_col=target_cell.column
                        )
                    ]
        if axis == 1:  # Tìm kiếm theo hàng ngang (row)
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == keyword:
                        target_cell = cell
                        break
                if target_cell:
                    cells = worksheet[target_cell.row]
                    break
        if not (target_cell and cells):
            self.logger.error(f"Không tìm thấy keyword: {keyword}")
            return None
        cells = [cell.value for cell in cells if cell.value is not None]
        values: list = cells[cells.index(keyword) + 1 :]
        values = " ".join(str(value) for value in values)
        self.logger.info(f"{self.title} - {keyword}: {values}")
        return values
